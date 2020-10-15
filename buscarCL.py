import openpyxl
def buscarCL(a):
    doc=openpyxl.load_workbook('instrucciones.xlsx')
    sheet=doc["LD"]
    for i in range(1,132):
        if sheet.cell(row=i,column=1).value==a:
            g=sheet.cell(row=i,column=3).value
            return g
