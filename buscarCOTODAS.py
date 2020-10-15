import openpyxl


def buscarCOTODAS(a):
    doc=openpyxl.load_workbook('instrucciones.xlsx')
    sheet=doc["TODAS"]
    for i in range(1,696):
        if(sheet.cell(row=i,column=1).value==a):
            h=sheet.cell(row=i,column=2).value
            return h
