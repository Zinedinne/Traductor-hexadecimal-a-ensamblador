import openpyxl

def buscarCLTODAS(a):
    doc=openpyxl.load_workbook('instrucciones.xlsx')

    sheet=doc["TODAS"]
    for i in range(1,696):
        if(sheet.cell(row=i,column=1).value==a):
                    g=sheet.cell(row=i,column=3).value
                    return g
