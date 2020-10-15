import openpyxl


def buscarCO(a):
    doc=openpyxl.load_workbook('instrucciones.xlsx')
    sheet=doc["LD"]
    for i in range(1,132):
        if(sheet.cell(row=i,column=1).value==a):
            h=sheet.cell(row=i,column=2).value
            return h
